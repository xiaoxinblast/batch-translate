"""xlsx 解析器：按行提取指定列"""

import json
import sys
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl，请执行 pip install openpyxl")
    sys.exit(1)


def parse(filepath: Path, source_col: str = "A", target_col: str = "B",
          header_row: int = 1, sheet_name: str | None = None, **opts) -> dict:
    """解析 xlsx，提取源列+上下文列。"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name == "*":
        worksheets = wb.worksheets
    elif sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"工作表不存在: {sheet_name}")
        worksheets = [wb[sheet_name]]
    else:
        worksheets = [wb.active]

    source_idx = _col_to_idx(source_col)
    target_idx = _col_to_idx(target_col) if target_col else None

    entries = []
    for ws in worksheets:
        max_col = max(ws.max_column or 0, source_idx + 1, (target_idx or 0) + 1)
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_col=max_col, values_only=True), 1
        ):
            if row_idx <= header_row:
                continue

            source = _cell(row, source_idx)
            if not source:
                continue
            target = _cell(row, target_idx) if target_idx is not None else ""

            # 搜集同行其他列作为 note
            note_parts = []
            for ci, cv in enumerate(row):
                if cv is not None and ci not in (source_idx, target_idx):
                    note_parts.append(str(cv).strip())
            note = " | ".join(note_parts) if note_parts else ""

            entries.append({
                "id": str(len(entries) + 1),
                "source": source,
                "target": target,
                "context": f"{filepath.name}:{ws.title}:Row{row_idx}",
                "note": note,
                "_sheet": ws.title,
                "_row": row_idx,
                "_col": source_idx,
                "_target_col": target_idx,
            })

    wb.close()
    return {
        "source_file": filepath.name,
        "header_row": header_row,
        "sheet_name": sheet_name,
        "entries": entries,
    }


def write(original_path: Path, translations_json: str | Path,
          output_path: Optional[Path] = None, **opts) -> Path:
    """将译文写回 xlsx 的目标列。"""
    with open(translations_json, "r", encoding="utf-8") as f:
        translations = json.load(f)

    target_map = {}
    row_map = {}
    col_map = {}
    sheet_map = {}
    if isinstance(translations, list):
        target_map = {str(r["id"]): r["target"] for r in translations}
    elif isinstance(translations, dict) and "entries" in translations:
        for e in translations["entries"]:
            if "target" in e and isinstance(e["target"], str):
                target_map[str(e["id"])] = e["target"]
            if e.get("_row") is not None:
                row_map[str(e["id"])] = int(e["_row"])
            if e.get("_target_col") is not None:
                col_map[str(e["id"])] = int(e["_target_col"])
            if e.get("_sheet"):
                sheet_map[str(e["id"])] = str(e["_sheet"])

    wb = openpyxl.load_workbook(
        original_path, keep_vba=original_path.suffix.lower() == ".xlsm"
    )

    updated = 0
    if row_map:
        # 优先使用 parse 阶段记录的 _row / _target_col（支持自定义表头行与目标列）
        for entry_id, target in target_map.items():
            row_idx = row_map.get(entry_id)
            if row_idx is None:
                continue
            col_idx = col_map.get(entry_id, 1) + 1  # 0-based → openpyxl 1-based；默认 B
            sheet = sheet_map.get(entry_id)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active
            ws.cell(row=row_idx, column=col_idx).value = target
            updated += 1
    else:
        # 纯数组输入回退：默认 B 列，行号减 1（表头行 1）
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            entry_id = str(row_idx - 1)
            if entry_id in target_map:
                ws.cell(row=row_idx, column=2).value = target_map[entry_id]
                updated += 1

    if output_path is None:
        output_path = original_path.with_stem(original_path.stem + "_translated")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()
    return output_path


def _col_to_idx(col: str) -> int:
    """A→0, B→1, ..."""
    col = col.upper().strip()
    result = 0
    for c in col:
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result - 1


def _cell(row: tuple, idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()
