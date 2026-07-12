#!/usr/bin/env python3
"""术语库：从 xlsx 加载术语，对日文原文做贪婪最长匹配。"""

import sys
from pathlib import Path
from typing import Optional

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl，请执行 pip install openpyxl")
    sys.exit(1)


class TermBase:
    """xlsx 术语库。列: 原文(ja) | 译文(zh) | 注释"""

    def __init__(self, xlsx_path: str | Path):
        self._path = Path(xlsx_path)
        self._terms: list[dict] = []  # [{ja, zh, note}]
        self._loaded = False

    # ── 加载 ──────────────────────────────────────────────────────

    def load(self) -> list[dict]:
        """加载术语表，返回术语列表。"""
        if self._loaded:
            return self._terms

        if not self._path.is_file():
            print(f"⚠️ 术语库不存在: {self._path}（将跳过术语匹配）", file=sys.stderr)
            self._loaded = True
            return self._terms

        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
        wb.close()

        for row in rows:
            ja = _cell(row, 0)
            zh = _cell(row, 1)
            if not ja:
                continue
            self._terms.append({
                "ja": ja,
                "zh": zh,
                "note": _cell(row, 2),
            })

        # 按原文长度降序排列，保证最长匹配优先
        self._terms.sort(key=lambda t: -len(t["ja"]))
        self._loaded = True
        return self._terms

    # ── 匹配 ──────────────────────────────────────────────────────

    def find_terms(self, text: str) -> list[dict]:
        """在 text 中查找所有匹配的术语，贪婪最长匹配。
        同源文位置允许多个不同译文共存。"""
        if not self._loaded:
            self.load()
        if not self._terms or not text:
            return []

        matched_spans: dict[tuple[int, str], list[dict]] = {}  # (pos, ja) → matches
        results: list[dict] = []

        for term in self._terms:
            ja = term["ja"]
            pos = 0
            while True:
                idx = text.find(ja, pos)
                if idx == -1:
                    break
                key = (idx, ja)
                term_range = set(range(idx, idx + len(ja)))
                # 只阻挡与"不同 ja"的已匹配区域重叠（避免短术语截断长术语）
                overlaps_diff = any(
                    other_ja != ja and set(range(oi, oi + len(other_ja))) & term_range
                    for oi, other_ja in matched_spans
                )
                if overlaps_diff:
                    pos = idx + 1
                    continue
                matched_spans.setdefault(key, []).append(
                    {"ja": ja, "zh": term["zh"], "note": term["note"]}
                )
                pos = idx + len(ja)

        for entries in matched_spans.values():
            results.extend(entries)
        return results


def _cell(row: tuple, col: int) -> str:
    if col >= len(row) or row[col] is None:
        return ""
    return str(row[col]).strip()


# ── 创建空白术语库模板 ──────────────────────────────────────────

def create_empty_term_base(path: str | Path) -> Path:
    """创建空白术语库 xlsx（含表头）。"""
    path = Path(path)
    if path.is_file():
        print(f"⚠️ 文件已存在，跳过: {path}")
        return path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "术语库"
    ws.append(["原文(ja)", "译文(zh)", "注释"])
    # 设置列宽
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()
    print(f"✅ 已创建空白术语库: {path}")
    return path


# ── CLI ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="术语库工具")
    p.add_argument("--create", type=str, help="创建空白术语库 xlsx")
    p.add_argument("--test", type=str, help="测试匹配（需指定术语库路径）")
    p.add_argument("--text", type=str, help="测试用日文原文")
    args = p.parse_args()

    if args.create:
        create_empty_term_base(args.create)
    elif args.test and args.text:
        tb = TermBase(args.test)
        tb.load()
        matches = tb.find_terms(args.text)
        for m in matches:
            print(f"  {m['ja']} → {m['zh']}  [{m['note']}]")
        if not matches:
            print("  (无匹配)")
    else:
        p.print_help()
