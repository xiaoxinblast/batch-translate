#!/usr/bin/env python3
"""用 translate 工作流的 parse 管线从已翻译批次重建 TM。

流程：
  1. mqxliff/docx/txt → convert.py parse → JSON
     xlsx/xlsm → openpyxl 直接读（parser 不提取 target 列）
  2. 从 entries 提取 source/target → TranslationMemory.add()
  3. 保存 TM
"""
import sys, json, subprocess, tempfile
import re
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SCRIPT_DIR))
from tm_store import TranslationMemory

XLSX_EXTS = {".xlsx", ".xlsm"}
PIPELINE_EXTS = {".mqxliff", ".docx", ".txt", ".csv", ".tsv"}


# ── xlsx/xlsm 直接提取 ────────────────────────────────────────────

def _col_idx(letter: str) -> int:
    """A→0, B→1, ..."""
    r = 0
    for c in letter.upper():
        r = r * 26 + (ord(c) - ord("A") + 1)
    return r - 1

def _cell(row: tuple, idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()

_SRC_KEYWORDS = ["原文", "source", "ja", "jp", "日文", "日本語"]
_TGT_KEYWORDS = ["译文", "target", "zh", "sc", "cn", "中文", "簡体", "简体"]
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
_TIME_RE = re.compile(r"^\d{2}:\d{2}")


def _header_scan(ws) -> tuple[int, list[int], list[int]]:
    """扫描前 5 行，返回 (header_row, 源候选列, 译文候选列)。
    header_row=0 表示未识别到表头。"""
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(5, ws.max_row or 5), values_only=True), 1
    ):
        src_cands = []
        tgt_cands = []
        for ci, cv in enumerate(row):
            cvs = str(cv).strip().lower() if cv else ""
            if any(kw in cvs for kw in _SRC_KEYWORDS):
                src_cands.append(ci)
            elif any(kw in cvs for kw in _TGT_KEYWORDS):
                tgt_cands.append(ci)
        if src_cands or tgt_cands:
            return row_idx, src_cands, tgt_cands
    return 0, [], []


def _pick_clean_source_col(ws, candidates: list[int], header_row: int) -> int | None:
    """在候选源列中选“干净文本列”：全表统计，注释/日期单元格计罚分，
    罚分最低者胜出；平局时数据行多者胜，再平局取最左列。无候选返回 None。"""
    if not candidates:
        return None
    start = header_row + 1 if header_row else 1
    scores = {ci: {"comment": 0, "datetime": 0, "nonempty": 0} for ci in candidates}
    for row_idx, row in enumerate(ws.iter_rows(min_row=start, values_only=True), start):
        for ci in candidates:
            v = _cell(row, ci)
            if not v:
                continue
            s = scores[ci]
            s["nonempty"] += 1
            if v.startswith("#") or v.startswith("//"):
                s["comment"] += 1
            elif _DATE_RE.match(v) or _TIME_RE.match(v):
                s["datetime"] += 1
    best = None
    best_penalty = None
    best_nonempty = None
    for ci in candidates:
        s = scores[ci]
        if s["nonempty"] == 0:
            penalty = 10 ** 9
        else:
            penalty = s["comment"] + s["datetime"]
        if best is None or penalty < best_penalty or (
            penalty == best_penalty and s["nonempty"] > best_nonempty
        ):
            best = ci
            best_penalty = penalty
            best_nonempty = s["nonempty"]
    return best


def _detect_columns(
    ws,
    source_col: str = "",
    target_col: str = "",
    header_row: int | None = None,
    clean_source: bool = False,
) -> tuple[int, int, int]:
    """检测源/译文列索引和表头行号。
    返回 (src_idx, tgt_idx, header_row)。header_row=0 表示无表头。
    - 显式列与自动列都会扫描前 5 行识别表头（除非 header_row 显式传入）
    - clean_source=True 时忽略 source_col，自动选无注释/日期的干净源列
    """
    detected_header, src_cands, tgt_cands = _header_scan(ws)
    if header_row is None:
        header_row = detected_header

    if clean_source:
        src_idx = _pick_clean_source_col(ws, src_cands, header_row)
        if src_idx is None:
            src_idx = _col_idx(source_col or "A")
    else:
        src_idx = _col_idx(source_col) if source_col else (src_cands[0] if src_cands else 0)

    if target_col:
        tgt_idx = _col_idx(target_col)
    else:
        tgt_idx = tgt_cands[0] if tgt_cands else (1 if src_idx != 1 else 2)
    if src_idx == tgt_idx:
        tgt_idx = 1 if src_idx != 1 else 2
    return src_idx, tgt_idx, header_row


def extract_xlsx(
    file_path: Path,
    source_col: str = "",
    target_col: str = "",
    header_row: int | None = None,
    clean_source: bool = False,
) -> list[dict]:
    """从 xlsx/xlsm 直接提取 source/target 对。"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    src_idx, tgt_idx, header_row = _detect_columns(
        ws, source_col, target_col, header_row=header_row, clean_source=clean_source
    )
    max_col = max(src_idx, tgt_idx) + 1

    entries = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), 1):
        if header_row and row_idx <= header_row:
            continue  # 跳过标题行及其之前的元数据行
        src = _cell(row, src_idx)
        tgt = _cell(row, tgt_idx)
        if not src or not tgt:
            continue
        entries.append({
            "source": src,
            "target": tgt,
            "context": f"{file_path.name}:Row{row_idx}",
            "file": file_path.name,
        })

    wb.close()
    return entries


# ── 通用 parse 管线（mqxliff/docx/txt 等） ────────────────────────

def parse_via_convert(file_path: Path, tmp_dir: Path) -> Path:
    """用 convert.py parse → JSON。"""
    output = tmp_dir / f"{file_path.stem}.json"
    subprocess.run([
        sys.executable, str(SCRIPT_DIR / "convert.py"), "parse",
        str(file_path),
        "--output", str(output),
        "--output-dir", str(tmp_dir),
    ], check=True, capture_output=True, text=True, encoding="utf-8")
    return output


# ── 主流程 ─────────────────────────────────────────────────────────

def main(
    src_dir: str,
    output: str,
    source_col: str = "",
    target_col: str = "",
    header_row: int | None = None,
    clean_source: bool = False,
):
    src = Path(src_dir)
    out = Path(output)
    if not src.is_dir():
        print(f"❌ 目录不存在: {src}")
        sys.exit(1)

    all_exts = list(PIPELINE_EXTS | XLSX_EXTS)
    files = sorted([f for ext in all_exts for f in src.glob(f"*{ext}")],
                   key=lambda f: f.name)
    if not files:
        print(f"❌ 目录中没有支持的文件: {src}")
        sys.exit(1)

    print(f"📂 找到 {len(files)} 个文件")

    tm = TranslationMemory(str(out.resolve()))
    tm._entries = []
    tm._loaded = True
    total_pairs = 0

    with tempfile.TemporaryDirectory(prefix="tm_rebuild_") as tmp_dir:
        tmp = Path(tmp_dir)
        for f in files:
            file_entries = []
            line_count = 0

            if f.suffix.lower() in XLSX_EXTS:
                # xlsx/xlsm：直接读
                file_entries = extract_xlsx(
                    f, source_col, target_col,
                    header_row=header_row, clean_source=clean_source,
                )
                line_count = len(file_entries)
            else:
                # mqxliff/docx/txt：走 convert.py parse 管线
                try:
                    json_path = parse_via_convert(f, tmp)
                except subprocess.CalledProcessError as e:
                    print(f"   ⚠️ {f.name}: parse 失败，跳过")
                    continue
                with open(json_path, "r", encoding="utf-8") as jf:
                    data = json.load(jf)
                entries = data.get("entries", [])
                line_count = len(entries)
                for e in entries:
                    src_text = (e.get("source") or "").strip()
                    tgt_text = (e.get("target") or "").strip()
                    if not src_text or not tgt_text:
                        continue
                    file_entries.append({
                        "source": src_text,
                        "target": tgt_text,
                        "context": e.get("context", "") or "",
                        "file": f.name,
                    })

            tm.add(file_entries, dedup=True)
            total_pairs += len(file_entries)
            status = f"{f.name}: {line_count} 条 → 提取 {len(file_entries)} 对"
            if len(file_entries) == 0 and f.suffix.lower() not in XLSX_EXTS:
                status += " ⚠️ 非 mqxliff 格式可能不提取译文，请用 --source-col/--target-col 或转为 mqxliff"
            print(f"   {status}")

    tm.save()
    size_mb = out.stat().st_size / (1024 * 1024)
    print(f"\n✅ TM 重建完成: {out}")
    print(f"   总提取: {total_pairs} 对")
    print(f"   去重后: {len(tm)} 条")
    print(f"   文件大小: {size_mb:.2f} MB")


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="从已翻译批次重建翻译记忆")
    p.add_argument("src_dir", help="已翻译批次目录")
    p.add_argument("--output", "-o", default="batch_translate/data/tm_memory.json",
                   help="输出 TM JSON 路径")
    p.add_argument("--source-col", default="", help="xlsx 源列字母（默认自动检测）")
    p.add_argument("--target-col", default="", help="xlsx 译文列字母（默认自动检测）")
    p.add_argument("--header-row", type=int, default=None,
                   help="xlsx 表头行号（默认自动检测；0=无表头）")
    p.add_argument("--clean-source", action="store_true",
                   help="自动选择无注释/日期的干净源列（忽略 --source-col）")
    args = p.parse_args()
    main(args.src_dir, args.output, args.source_col, args.target_col,
         header_row=args.header_row, clean_source=args.clean_source)
